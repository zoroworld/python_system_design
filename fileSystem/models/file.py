from abc import ABC, abstractmethod
import os
from reportlab.pdfgen import canvas
import openpyxl

class File(ABC):

    @abstractmethod
    def create_file(self):
        pass

    @abstractmethod
    def delete_file(self):
        pass

    @abstractmethod
    def update_file(self):
        pass

    @abstractmethod
    def read_file(self):
        pass


class PdfFile:
    def __init__(self, name):
        self.folder = "./pdffolder"
        os.makedirs(self.folder, exist_ok=True)  # ensure folder exists
        self.name = name if name.endswith(".pdf") else name + ".pdf"
        self.filepath = os.path.join(self.folder, self.name)
        self.create_file()

    def create_file(self):
        if not os.path.exists(self.filepath):
            c = canvas.Canvas(self.filepath)
            c.drawString(100, 750, f"{self.name} created!")  # placeholder text
            c.save()
            print(f"{self.name} file created in {self.folder}")
        else:
            print(f"{self.name} already exists!")

    def delete_file(self):
        if os.path.exists(self.filepath):
            os.remove(self.filepath)
            print(f"{self.name} deleted.")
        else:
            print(f"{self.name} does not exist.")

    def update_file(self, text):
        if os.path.exists(self.filepath):
            c = canvas.Canvas(self.filepath)
            c.drawString(100, 750, text)  # overwrite with new text
            c.save()
            print(f"{self.name} updated with new content.")
        else:
            print(f"{self.name} does not exist.")

    def read_file(self):
        # PDF isn't plain text, so we can't just open() it.
        # If you want to extract text, you can use PyPDF2 or pdfplumber
        if os.path.exists(self.filepath):
            print(f"{self.name} exists at {self.filepath} (cannot read raw text directly).")
        else:
            print(f"{self.name} does not exist.")




class TxtFile:
    def __init__(self, name):
        self.folder = "./txtfolder"
        # ensure folder exists
        if not os.path.exists(self.folder):
            os.makedirs(self.folder)
        self.name = name

    def create_file(self):
        store = os.path.join(self.folder, self.name)
        try:
            with open(store, "x") as f:
                pass
            print(f"{self.name} file is created !!!")
        except FileExistsError:
            print(f"{self.name} already exists!")

    def delete_file(self):
        store = os.path.join(self.folder, self.name)
        if os.path.exists(store):
            os.remove(store)
            print(f"{self.name} deleted.")
        else:
            print(f"{self.name} does not exist.")

    def update_file(self, content):
        store = os.path.join(self.folder, self.name)
        with open(store, "a") as f:  # append mode
            f.write(content + "\n")
        print(f"{self.name} updated with new content.")

    def read_file(self):
        store = os.path.join(self.folder, self.name)
        if os.path.exists(store):
            with open(store, "r") as f:
                data = f.read()
            print(data)
            return data
        else:
            print(f"{self.name} does not exist.")
            return None

class XlsFile:
    def __init__(self, name):
        self.folder = "./xlsfolder"
        os.makedirs(self.folder, exist_ok=True)  # ensure folder exists
        self.name = name if name.endswith(".xlsx") else name + ".xlsx"
        self.filepath = os.path.join(self.folder, self.name)
        self.create_file()

    def create_file(self):
        if not os.path.exists(self.filepath):
            wb = openpyxl.Workbook()
            wb.save(self.filepath)
            print(f"{self.name} file created in {self.folder}")
        else:
            print(f"{self.name} already exists!")

    def delete_file(self):
        if os.path.exists(self.filepath):
            os.remove(self.filepath)
            print(f"{self.name} deleted.")
        else:
            print(f"{self.name} does not exist.")

    def update_file(self, sheet="Sheet1", cell="A1", value=""):
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb[sheet]
        ws[cell] = value
        wb.save(self.filepath)
        print(f"{self.name} updated: {cell} = {value}")

    def read_file(self, sheet="Sheet1"):
        if os.path.exists(self.filepath):
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb[sheet]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            print(f"Content of {self.name}:")
            for r in data:
                print(r)
            return data
        else:
            print(f"{self.name} does not exist.")
            return None
