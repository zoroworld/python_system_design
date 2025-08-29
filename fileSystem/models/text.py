from abc import ABC, abstractmethod
import os
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader
import openpyxl

class Text(ABC):

    @abstractmethod
    def appedText(self):
        pass
    @abstractmethod
    def deleteText(self):
        pass
    @abstractmethod
    def addText(self):
        pass
    @abstractmethod
    def readText(self):
        pass



class PdfText:
    def __init__(self, content, text, sign, file):
        self.content = content
        self.text = text
        self.sign = sign
        self.file = file  # should be a PdfFile instance

    def appedText(self, content=None, text=None, sign=None):
        """Append text to the PDF (new page)."""
        c = canvas.Canvas(self.file.filepath)
        c.drawString(100, 750, content or self.content)
        c.drawString(100, 730, text or self.text)
        c.drawString(100, 710, sign or self.sign)
        c.save()
        print(f"Appended text to {self.file.name}")

    def deleteText(self, *args):
        """Clear the PDF (overwrite with empty)."""
        c = canvas.Canvas(self.file.filepath)
        c.save()
        print(f"Cleared all text in {self.file.name}")

    def addText(self, content=None, text=None, sign=None):
        """Overwrite the PDF with new text."""
        c = canvas.Canvas(self.file.filepath)
        c.drawString(100, 750, content or self.content)
        c.drawString(100, 730, text or self.text)
        c.drawString(100, 710, sign or self.sign)
        c.save()
        print(f"Added new text to {self.file.name}")

    def readText(self, *args):
        """Read text from the PDF (extract using PyPDF2)."""
        if os.path.exists(self.file.filepath):
            reader = PdfReader(self.file.filepath)
            text_data = ""
            for page in reader.pages:
                text_data += page.extract_text() or ""
            print("PDF Content:\n", text_data)
            return text_data
        else:
            print(f"{self.file.name} does not exist.")
            return None




class TxtfText:
    def __init__(self, text, file):
        self.text = text
        self.file = file  # file should be a TexFile instance

    def appedText(self):
        """Append text to the file."""
        store = os.path.join(self.file.folder, self.file.name)
        with open(store, "a") as f:
            f.write(self.text + "\n")
        print(f"Appended text to {self.file.name}")

    def deleteText(self):
        """Delete all text (clear the file)."""
        store = os.path.join(self.file.folder, self.file.name)
        with open(store, "w") as f:  # overwrite with nothing
            pass
        print(f"All text deleted from {self.file.name}")

    def addText(self):
        """Overwrite the file with new text."""
        store = os.path.join(self.file.folder, self.file.name)
        with open(store, "w") as f:
            f.write(self.text + "\n")
        print(f"Added (overwritten) text to {self.file.name}")

    def readText(self):
        """Read text from the file."""
        store = os.path.join(self.file.folder, self.file.name)
        if os.path.exists(store):
            with open(store, "r") as f:
                data = f.read()
            print(data)
            return data
        else:
            print(f"{self.file.name} does not exist.")
            return None



class XlsText:
    def __init__(self, file, sheet="Sheet1"):
        self.file = file  # should be an XlsFile instance
        self.sheet = sheet

    def appendText(self, value):
        wb = openpyxl.load_workbook(self.file.filepath)
        ws = wb[self.sheet]
        ws.append([value])
        wb.save(self.file.filepath)
        print(f"Appended '{value}' to {self.file.name}")

    def addText(self, cell, value):
        wb = openpyxl.load_workbook(self.file.filepath)
        ws = wb[self.sheet]
        ws[cell] = value
        wb.save(self.file.filepath)
        print(f"Added '{value}' at {cell} in {self.file.name}")

    def deleteText(self, cell):
        wb = openpyxl.load_workbook(self.file.filepath)
        ws = wb[self.sheet]
        ws[cell] = None
        wb.save(self.file.filepath)
        print(f"Deleted text from {cell} in {self.file.name}")

    def readText(self):
        wb = openpyxl.load_workbook(self.file.filepath)
        ws = wb[self.sheet]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        print(f"Data in {self.file.name}:")
        for r in data:
            print(r)
        return data
